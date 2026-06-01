from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from datetime import datetime


# =============================================================================
# MAPA DE NIVEIS HIERARQUICOS
# =============================================================================
NIVEIS = {
    1: 'Técnico',
    2: 'Supervisor',
    3: 'Chefe de Departamento',
    4: 'Subdirector',
    5: 'Director',
    6: 'Administrador'
}


# =============================================================================
# EXPORTACAO PDF
# =============================================================================

def gerar_pdf(auditoria):
    """
    Gera um relatorio PDF completo de uma sessao de auditoria.
    Devolve os bytes do ficheiro PDF.
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=2*cm,
        leftMargin=2*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )

    styles = getSampleStyleSheet()
    elementos = []

    # Estilos personalizados
    estilo_titulo = ParagraphStyle(
        'titulo',
        parent=styles['Title'],
        fontSize=16,
        textColor=colors.HexColor('#0d6efd'),
        alignment=TA_CENTER,
        spaceAfter=6
    )

    estilo_subtitulo = ParagraphStyle(
        'subtitulo',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.grey,
        alignment=TA_CENTER,
        spaceAfter=20
    )

    estilo_secao = ParagraphStyle(
        'secao',
        parent=styles['Heading2'],
        fontSize=12,
        textColor=colors.HexColor('#0d6efd'),
        spaceBefore=16,
        spaceAfter=8
    )

    estilo_normal = ParagraphStyle(
        'normal_custom',
        parent=styles['Normal'],
        fontSize=9,
        leading=12
    )

    # Cabecalho
    elementos.append(Paragraph('NS Aplicação', estilo_titulo))
    elementos.append(
        Paragraph('Sistema de Auditoria de Conformidade', estilo_subtitulo))
    elementos.append(Spacer(1, 0.3*cm))

    # Dados da sessao
    elementos.append(Paragraph('Relatório de Auditoria', estilo_secao))

    dados_sessao = [
        ['Campo', 'Valor'],
        ['ID da Auditoria', f'#{auditoria.id_auditoria}'],
        ['Período Analisado', auditoria.periodo_analisado],
        ['Data de Execução', auditoria.data_execucao.strftime(
            '%d/%m/%Y %H:%M')],
        ['Auditor', auditoria.utilizador.nome],
        ['Status', auditoria.status.upper()],
        ['Total de Aquisições', str(auditoria.total_transacoes)],
        ['Não Conformidades', str(auditoria.total_nao_conformidades)],
    ]

    tabela_sessao = Table(dados_sessao, colWidths=[6*cm, 11*cm])
    tabela_sessao.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d6efd')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#f8f9fa')),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1),
         [colors.white, colors.HexColor('#f8f9fa')]),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    elementos.append(tabela_sessao)
    elementos.append(Spacer(1, 0.5*cm))

    # Resultados por aquisicao
    elementos.append(Paragraph('Resultados por Aquisição', estilo_secao))

    cabecalho_aquisicoes = ['#', 'Descrição', 'Valor (Kz)', 'Resultado']
    linhas_aquisicoes = [cabecalho_aquisicoes]

    for aa in auditoria.avaliacoes:
        resultado = 'CONFORME' if aa.resultado == 'conforme' else 'NÃO CONFORME'
        linhas_aquisicoes.append([
            str(aa.aquisicao.id_aquisicao),
            Paragraph(aa.aquisicao.descricao[:60], estilo_normal),
            f"{aa.aquisicao.valor:,.2f}",
            resultado
        ])

    tabela_aquisicoes = Table(
        linhas_aquisicoes,
        colWidths=[1.5*cm, 8*cm, 4*cm, 3.5*cm]
    )
    tabela_aquisicoes.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d6efd')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1),
         [colors.white, colors.HexColor('#f8f9fa')]),
        ('PADDING', (0, 0), (-1, -1), 5),
        ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
        ('ALIGN', (3, 0), (3, -1), 'CENTER'),
    ]))

    # Colorir linhas nao conformes
    for i, aa in enumerate(auditoria.avaliacoes, start=1):
        if aa.resultado == 'nao_conforme':
            tabela_aquisicoes.setStyle(TableStyle([
                ('TEXTCOLOR', (3, i), (3, i), colors.HexColor('#dc3545')),
                ('FONTNAME', (3, i), (3, i), 'Helvetica-Bold'),
            ]))

    elementos.append(tabela_aquisicoes)
    elementos.append(Spacer(1, 0.5*cm))

    # Nao conformidades
    nao_conformidades = []
    for aa in auditoria.avaliacoes:
        for nc in aa.nao_conformidades:
            nao_conformidades.append(nc)

    if nao_conformidades:
        elementos.append(
            Paragraph('Não Conformidades Detectadas', estilo_secao))

        cabecalho_nc = ['#', 'Aquisição', 'Regra',
                        'Descrição', 'Gravidade', 'Status']
        linhas_nc = [cabecalho_nc]

        for nc in nao_conformidades:
            linhas_nc.append([
                str(nc.id_nao_conformidade),
                f"#{nc.auditoria_aquisicao.aquisicao.id_aquisicao}",
                nc.regra.codigo,
                Paragraph(nc.descricao[:80], estilo_normal),
                nc.gravidade.upper(),
                nc.status.upper()
            ])

        tabela_nc = Table(
            linhas_nc,
            colWidths=[1*cm, 2*cm, 1.5*cm, 8*cm, 2*cm, 2.5*cm]
        )
        tabela_nc.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dc3545')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1),
             [colors.white, colors.HexColor('#fff3f3')]),
            ('PADDING', (0, 0), (-1, -1), 4),
            ('ALIGN', (4, 0), (4, -1), 'CENTER'),
            ('ALIGN', (5, 0), (5, -1), 'CENTER'),
        ]))
        elementos.append(tabela_nc)

    # Rodape
    elementos.append(Spacer(1, 1*cm))
    elementos.append(Paragraph(
        f'Relatório gerado em {datetime.now().strftime("%d/%m/%Y %H:%M")} — NS Aplicação © 2025',
        ParagraphStyle('rodape', parent=styles['Normal'], fontSize=7,
                       textColor=colors.grey, alignment=TA_CENTER)
    ))

    doc.build(elementos)
    buffer.seek(0)
    return buffer


# =============================================================================
# EXPORTACAO EXCEL
# =============================================================================

def gerar_excel(auditoria):
    """
    Gera um ficheiro Excel com três folhas:
    - Resumo da auditoria
    - Resultados por aquisicao
    - Nao conformidades detalhadas
    Devolve os bytes do ficheiro Excel.
    """
    wb = Workbook()

    # Estilos
    fonte_titulo = Font(name='Calibri', bold=True, size=14, color='0D6EFD')
    fonte_cabecalho = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
    fonte_normal = Font(name='Calibri', size=10)
    fonte_bold = Font(name='Calibri', bold=True, size=10)

    fill_azul = PatternFill(start_color='0D6EFD',
                            end_color='0D6EFD', fill_type='solid')
    fill_vermelho = PatternFill(
        start_color='DC3545', end_color='DC3545', fill_type='solid')
    fill_cinza = PatternFill(start_color='F8F9FA',
                             end_color='F8F9FA', fill_type='solid')
    fill_verde = PatternFill(start_color='D4EDDA',
                             end_color='D4EDDA', fill_type='solid')
    fill_rosa = PatternFill(start_color='F8D7DA',
                            end_color='F8D7DA', fill_type='solid')

    borda = Border(
        left=Side(style='thin', color='DEE2E6'),
        right=Side(style='thin', color='DEE2E6'),
        top=Side(style='thin', color='DEE2E6'),
        bottom=Side(style='thin', color='DEE2E6')
    )

    alinhamento_centro = Alignment(horizontal='center', vertical='center')
    alinhamento_esquerda = Alignment(
        horizontal='left', vertical='center', wrap_text=True)

    def aplicar_cabecalho(ws, cabecalhos, linha=1, fill=None):
        """Aplica estilo de cabecalho a uma linha."""
        fill_usar = fill or fill_azul
        for col, texto in enumerate(cabecalhos, start=1):
            cell = ws.cell(row=linha, column=col, value=texto)
            cell.font = fonte_cabecalho
            cell.fill = fill_usar
            cell.alignment = alinhamento_centro
            cell.border = borda

    def aplicar_borda_linha(ws, linha, num_colunas):
        """Aplica borda a todas as células de uma linha."""
        for col in range(1, num_colunas + 1):
            ws.cell(row=linha, column=col).border = borda

    # =========================================================================
    # FOLHA 1 — RESUMO
    # =========================================================================
    ws_resumo = wb.active
    ws_resumo.title = 'Resumo'

    # Titulo
    ws_resumo.merge_cells('A1:C1')
    cell_titulo = ws_resumo['A1']
    cell_titulo.value = 'NS Aplicação — Relatório de Auditoria de Conformidade'
    cell_titulo.font = fonte_titulo
    cell_titulo.alignment = alinhamento_centro
    ws_resumo.row_dimensions[1].height = 30

    ws_resumo.merge_cells('A2:C2')
    cell_sub = ws_resumo['A2']
    cell_sub.value = f'Período: {auditoria.periodo_analisado} | Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    cell_sub.font = Font(name='Calibri', size=9, color='6C757D', italic=True)
    cell_sub.alignment = alinhamento_centro

    ws_resumo.append([])

    # Dados do resumo
    aplicar_cabecalho(ws_resumo, ['Campo', 'Valor', ''], linha=4)

    dados_resumo = [
        ['ID da Auditoria', f'#{auditoria.id_auditoria}', ''],
        ['Período Analisado', auditoria.periodo_analisado, ''],
        ['Data de Execução', auditoria.data_execucao.strftime(
            '%d/%m/%Y %H:%M'), ''],
        ['Auditor Responsável', auditoria.utilizador.nome, ''],
        ['Perfil do Auditor', auditoria.utilizador.perfil.title(), ''],
        ['Status da Auditoria', auditoria.status.upper(), ''],
        ['Total de Aquisições Analisadas', auditoria.total_transacoes, ''],
        ['Aquisições Conformes', auditoria.total_transacoes -
            auditoria.total_nao_conformidades, ''],
        ['Aquisições Não Conformes', auditoria.total_nao_conformidades, ''],
    ]

    for i, linha in enumerate(dados_resumo, start=5):
        ws_resumo.cell(row=i, column=1, value=linha[0]).font = fonte_bold
        ws_resumo.cell(row=i, column=1).fill = fill_cinza
        ws_resumo.cell(row=i, column=2, value=linha[1]).font = fonte_normal
        aplicar_borda_linha(ws_resumo, i, 3)

        # Colorir conformes e nao conformes
        if 'Não Conformes' in linha[0]:
            ws_resumo.cell(row=i, column=2).fill = fill_rosa
            ws_resumo.cell(row=i, column=2).font = Font(name='Calibri', size=10,
                                                        bold=True, color='DC3545')
        elif 'Conformes' in linha[0] and 'Não' not in linha[0]:
            ws_resumo.cell(row=i, column=2).fill = fill_verde
            ws_resumo.cell(row=i, column=2).font = Font(name='Calibri', size=10,
                                                        bold=True, color='198754')

    ws_resumo.column_dimensions['A'].width = 35
    ws_resumo.column_dimensions['B'].width = 30
    ws_resumo.column_dimensions['C'].width = 10

    # =========================================================================
    # FOLHA 2 — RESULTADOS POR AQUISICAO
    # =========================================================================
    ws_aquisicoes = wb.create_sheet('Resultados por Aquisição')

    aplicar_cabecalho(ws_aquisicoes, [
        '#', 'Descrição', 'Centro de Custo', 'Valor (Kz)',
        'Data Solicitação', 'Data Aprovação', 'Status Aprovação',
        'Tipo', 'Resultado Auditoria'
    ])

    for i, aa in enumerate(auditoria.avaliacoes, start=2):
        a = aa.aquisicao
        ws_aquisicoes.cell(row=i, column=1, value=a.id_aquisicao)
        ws_aquisicoes.cell(row=i, column=2, value=a.descricao)
        ws_aquisicoes.cell(row=i, column=3, value=a.centro.nome)
        ws_aquisicoes.cell(row=i, column=4, value=float(a.valor))
        ws_aquisicoes.cell(row=i, column=4).number_format = '#,##0.00'
        ws_aquisicoes.cell(
            row=i, column=5, value=a.data_solicitacao.strftime('%d/%m/%Y'))
        ws_aquisicoes.cell(row=i, column=6, value=a.data_aprovacao.strftime(
            '%d/%m/%Y') if a.data_aprovacao else '—')
        ws_aquisicoes.cell(row=i, column=7, value=a.status_aprovacao.title())
        ws_aquisicoes.cell(row=i, column=8, value=a.tipo_aquisicao.title())
        ws_aquisicoes.cell(
            row=i, column=9, value='CONFORME' if aa.resultado == 'conforme' else 'NÃO CONFORME')

        # Colorir resultado
        cell_resultado = ws_aquisicoes.cell(row=i, column=9)
        if aa.resultado == 'conforme':
            cell_resultado.fill = fill_verde
            cell_resultado.font = Font(
                name='Calibri', size=10, bold=True, color='198754')
        else:
            cell_resultado.fill = fill_rosa
            cell_resultado.font = Font(
                name='Calibri', size=10, bold=True, color='DC3545')

        aplicar_borda_linha(ws_aquisicoes, i, 9)
        for col in range(1, 10):
            ws_aquisicoes.cell(row=i, column=col).font = fonte_normal
            ws_aquisicoes.cell(
                row=i, column=col).alignment = alinhamento_esquerda

    ws_aquisicoes.column_dimensions['A'].width = 8
    ws_aquisicoes.column_dimensions['B'].width = 40
    ws_aquisicoes.column_dimensions['C'].width = 30
    ws_aquisicoes.column_dimensions['D'].width = 18
    ws_aquisicoes.column_dimensions['E'].width = 18
    ws_aquisicoes.column_dimensions['F'].width = 18
    ws_aquisicoes.column_dimensions['G'].width = 18
    ws_aquisicoes.column_dimensions['H'].width = 12
    ws_aquisicoes.column_dimensions['I'].width = 18

    # =========================================================================
    # FOLHA 3 — NAO CONFORMIDADES
    # =========================================================================
    ws_nc = wb.create_sheet('Não Conformidades')

    aplicar_cabecalho(ws_nc, [
        '#', 'Aquisição', 'Regra', 'Nome da Regra',
        'Descrição da Irregularidade', 'Gravidade', 'Data Registo',
        'Status', 'Comentário do Auditor'
    ], fill=fill_vermelho)

    linha_nc = 2
    for aa in auditoria.avaliacoes:
        for nc in aa.nao_conformidades:
            ws_nc.cell(row=linha_nc, column=1, value=nc.id_nao_conformidade)
            ws_nc.cell(row=linha_nc, column=2,
                       value=f"#{aa.aquisicao.id_aquisicao}")
            ws_nc.cell(row=linha_nc, column=3, value=nc.regra.codigo)
            ws_nc.cell(row=linha_nc, column=4, value=nc.regra.nome)
            ws_nc.cell(row=linha_nc, column=5, value=nc.descricao)
            ws_nc.cell(row=linha_nc, column=6, value=nc.gravidade.upper())
            ws_nc.cell(row=linha_nc, column=7,
                       value=nc.data_registo.strftime('%d/%m/%Y'))
            ws_nc.cell(row=linha_nc, column=8, value=nc.status.upper())
            ws_nc.cell(row=linha_nc, column=9,
                       value=nc.comentario_auditor or '—')

            # Colorir gravidade
            cell_gravidade = ws_nc.cell(row=linha_nc, column=6)
            if nc.gravidade == 'critica':
                cell_gravidade.fill = PatternFill(start_color='DC3545',
                                                  end_color='DC3545', fill_type='solid')
                cell_gravidade.font = Font(name='Calibri', size=10,
                                           bold=True, color='FFFFFF')
            elif nc.gravidade == 'alta':
                cell_gravidade.fill = PatternFill(start_color='FD7E14',
                                                  end_color='FD7E14', fill_type='solid')
                cell_gravidade.font = Font(name='Calibri', size=10,
                                           bold=True, color='FFFFFF')
            elif nc.gravidade == 'media':
                cell_gravidade.fill = PatternFill(start_color='FFC107',
                                                  end_color='FFC107', fill_type='solid')
                cell_gravidade.font = Font(name='Calibri', size=10, bold=True)

            aplicar_borda_linha(ws_nc, linha_nc, 9)
            for col in range(1, 10):
                if col != 6:
                    ws_nc.cell(row=linha_nc, column=col).font = fonte_normal
                ws_nc.cell(
                    row=linha_nc, column=col).alignment = alinhamento_esquerda

            linha_nc += 1

    ws_nc.column_dimensions['A'].width = 8
    ws_nc.column_dimensions['B'].width = 12
    ws_nc.column_dimensions['C'].width = 10
    ws_nc.column_dimensions['D'].width = 30
    ws_nc.column_dimensions['E'].width = 50
    ws_nc.column_dimensions['F'].width = 12
    ws_nc.column_dimensions['G'].width = 15
    ws_nc.column_dimensions['H'].width = 15
    ws_nc.column_dimensions['I'].width = 40

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
